import re
import sys
import json
import asyncio
import signal
import datetime
import urllib.parse
from collections import defaultdict
from pathlib import Path

if sys.platform == "win32":
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
except ImportError:
    print("Chưa cài Playwright!")
    print("pip install playwright")
    print("playwright install chromium")
    sys.exit(1)

SITEMAP_INDEX = "https://www.tequipment.net/sitemap-index3.xml"
BASE_URL      = "https://www.tequipment.net"
OUTPUT_FILE   = "tequipment_pricelists.xlsx"
CHECKPOINT    = "checkpoint.json"

MAX_PRODUCTS  = 100
WORKERS       = 3
DELAY_SECS    = 1.5
HEADLESS      = True
TIMEOUT_MS    = 30_000

stop_flag = False
counter   = {"done": 0, "errors": 0, "blocked": 0}
done_urls = set()


def parse_locs(text: str) -> list:
    return re.findall(r"<loc>\s*(https?://[^\s<]+)\s*</loc>", text)

DISALLOWED = ["/search/", "/admin/", "/friend/", "/shopping-cart/",
              "/assets/TaxExemptDocs/", "/assets/PODocuments/"]

def is_allowed(url: str) -> bool:
    path = urllib.parse.urlparse(url).path
    return not any(path.startswith(d) for d in DISALLOWED)

def is_product_url(url: str) -> bool:
    if url.endswith(".html"):
        return True
    parts = [p for p in urllib.parse.urlparse(url).path.strip("/").split("/") if p]
    if len(parts) < 2:
        return False
    skip = {"assets","my-account","shopping-cart","clearance","promotions",
            "brands","departments","request-quote","shipping","payment",
            "sitemaps","search","admin","open-box-b-stock","nist",
            "interactive-technology"}
    return parts[0].lower() not in skip

def is_incapsula(html: str) -> bool:
    return "_Incapsula_Resource" in html or "incident_id" in html


def load_checkpoint() -> set:
    if not Path(CHECKPOINT).exists():
        return set()
    try:
        with open(CHECKPOINT, "r", encoding="utf-8") as f:
            data = json.load(f)
        done = set(data.get("done_urls", []))
        print(f"  [Resume] {len(done):,} URLs đã crawl từ trước")
        return done
    except Exception:
        return set()

def save_checkpoint():
    try:
        with open(CHECKPOINT, "w", encoding="utf-8") as f:
            json.dump({"done_urls": list(done_urls),
                       "timestamp": datetime.datetime.now().isoformat()}, f)
    except Exception as e:
        print(f"  [WARN] Checkpoint: {e}")


def extract_brand_model_cat(url: str) -> tuple:
    if url.endswith(".html"):
        stem = Path(urllib.parse.urlparse(url).path).stem
        m = re.match(r"([A-Za-z][A-Za-z\-]*)(.+)", stem)
        return (m.group(1), m.group(2), "") if m else (stem, "", "")
    parts = [p for p in urllib.parse.urlparse(url).path.strip("/").split("/") if p]
    brand = parts[0] if parts else ""
    model = parts[1] if len(parts) > 1 else ""
    cat   = parts[2].replace("-", " ").title() if len(parts) > 2 else ""
    return brand, model, cat

def parse_product(url: str, html: str) -> dict:
    soup  = BeautifulSoup(html, "lxml")
    
    brand, model, cat_url = extract_brand_model_cat(url)
    h1   = soup.find("h1")
    name = h1.get_text(strip=True) if h1 else f"{brand} {model}"
    crumbs = []
    ol = soup.find("ol")
    if ol:
        for li in ol.find_all("li"):
            t = li.get_text(strip=True)
            if t and t.lower() not in ("home", ""):
                crumbs.append(t)
    if len(crumbs) > 2:
        category = " > ".join(crumbs[:-2])
    elif len(crumbs) == 2:
        category = crumbs[0]
    else:
        category = cat_url

    page_text  = soup.get_text(" ", strip=True)
    list_price = ""
    your_price = ""
    in_stock   = ""

    m = re.search(r"List\s+Price\s*:\s*\$\s*([\d,]+\.?\d*)", page_text, re.I)
    if m:
        list_price = "$" + m.group(1)

    m = re.search(r"Your\s+Price\s*:\s*\$\s*([\d,]+\.?\d*)", page_text, re.I)
    if m:
        your_price = "$" + m.group(1)

    if not your_price:
        m = re.search(r"(?:Sale|Our)\s+Price\s*:\s*\$\s*([\d,]+\.?\d*)", page_text, re.I)
        if m:
            your_price = "$" + m.group(1)

    if not list_price:
        for script in soup.find_all("script", type="application/ld+json"):
            try:
                data   = json.loads(script.string or "")
                offers = data.get("offers", {})
                if isinstance(offers, list):
                    offers = offers[0]
                price = str(offers.get("price", ""))
                if price and price not in ("0", ""):
                    list_price = "$" + price
                    break
            except Exception:
                pass

    if not list_price and not your_price:
        if re.search(r"Let us quote|quote you our best price", page_text, re.I):
            list_price = "Quote only"

    m = re.search(r"In\s+Stock\s*:\s*([^\n\r]{1,60})", page_text, re.I)
    if m:
        raw = re.sub(r"\s*(Free\s+shipping|View\s+Payment|Add\s+to|Request|Sign).*",
                     "", m.group(1), flags=re.I).strip()
        in_stock = raw[:50]

    description = ""
    meta_d = soup.find("meta", {"name": "description"})
    if meta_d and meta_d.get("content"):
        description = meta_d["content"].strip()[:200]

    return {"brand": brand, "model": model, "name": name, "category": category,
            "list_price": list_price, "your_price": your_price,
            "in_stock": in_stock, "description": description, "url": url}


async def safe_goto(page, url: str, retries: int = 2) -> str | None:
    for attempt in range(retries + 1):
        try:
            await page.goto(url, timeout=TIMEOUT_MS, wait_until="networkidle")
            await asyncio.sleep(DELAY_SECS)
            html = await page.content()
            if not is_incapsula(html):
                return html
            print(f"  [Incapsula attempt {attempt+1}] {url}")
            await page.goto(BASE_URL, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
            await asyncio.sleep(3)
        except PWTimeout:
            print(f"  [Timeout attempt {attempt+1}] {url}")
            await asyncio.sleep(2)
        except Exception as e:
            print(f"  [Error] {url}: {type(e).__name__}: {e}")
            await asyncio.sleep(2)
    return None


async def worker(wid: int, queue: asyncio.Queue, results: list, context) -> None:
    page = await context.new_page()
    try:
        await page.goto(BASE_URL, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
        await asyncio.sleep(2)
    except Exception:
        pass

    while True:
        if stop_flag:
            break
        try:
            url = queue.get_nowait()
        except asyncio.QueueEmpty:
            break

        html = await safe_goto(page, url)
        if html is None:
            counter["errors"] += 1
        elif is_incapsula(html):
            counter["blocked"] += 1
        else:
            try:
                product = parse_product(url, html)
                results.append(product)
                done_urls.add(url)
                counter["done"] += 1
                n  = counter["done"]
                lp = product["list_price"] or "—"
                yp = product["your_price"] or "—"
                if n <= 30 or n % 100 == 0:
                    print(f"    [{n:>5}] {product['brand']:<18} {product['model']:<22} "
                          f"List:{lp:<10} Sale:{yp}")
                if n % 500 == 0:
                    save_checkpoint()
            except Exception as e:
                counter["errors"] += 1
                print(f"    [PARSE ERR] {url}: {e}")

        queue.task_done()

    await page.close()


async def get_sitemap_urls(context) -> list:
    print("[STEP 1] Reading sitemaps...")
    page = await context.new_page()
    await page.goto(BASE_URL, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
    await asyncio.sleep(2)

    print(f"    Fetching: {SITEMAP_INDEX}")
    await page.goto(SITEMAP_INDEX, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
    child_sitemaps = parse_locs(await page.content())

    if not child_sitemaps:
        child_sitemaps = [
            f"{BASE_URL}/sitemaps/sitemap3_items{i}.xml" for i in range(1, 10)
        ]
        print(f"    [WARN] Dùng fallback list")
    else:
        print(f"    → {len(child_sitemaps)} child sitemaps")

    product_urls = []
    for sm in [u for u in child_sitemaps if "items" in u]:
        print(f"    Reading: {sm}")
        await page.goto(sm, timeout=TIMEOUT_MS, wait_until="domcontentloaded")
        urls  = parse_locs(await page.content())
        valid = [u for u in urls if is_allowed(u) and is_product_url(u)]
        product_urls.extend(valid)
        print(f"    → {len(valid):,} URLs")
        await asyncio.sleep(0.5)

    await page.close()
    product_urls = list(dict.fromkeys(product_urls))
    print(f"\n    Total: {len(product_urls):,} product URLs\n")
    return product_urls


async def main_async():
    global stop_flag

    print("=" * 65)
    print("  TEquipment.net Price List Crawler")
    print("=" * 65 + "\n")

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=HEADLESS)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 800},
            locale="en-US",
        )

        product_urls = await get_sitemap_urls(context)

        already_done = load_checkpoint()
        remaining    = [u for u in product_urls if u not in already_done]
        print(f"    Remaining: {len(remaining):,} (skip {len(already_done):,})")

        if MAX_PRODUCTS > 0:
            remaining = remaining[:MAX_PRODUCTS]
            print(f"    [TEST MODE] Giới hạn {MAX_PRODUCTS} products")

        total = len(remaining)
        print(f"\n[STEP 2] Crawling {total:,} products ({WORKERS} tabs)...\n")
        print(f"  {'#':>5}  {'Brand':<18} {'Model':<22} {'List Price':<12} Sale Price")
        print(f"  {'─'*72}")

        queue = asyncio.Queue()
        for url in remaining:
            await queue.put(url)

        results  = []
        n_workers = min(WORKERS, total) if total > 0 else 1
        tasks    = [
            asyncio.create_task(worker(i+1, queue, results, context))
            for i in range(n_workers)
        ]
        await asyncio.gather(*tasks)

        await context.close()
        await browser.close()

    save_checkpoint()

    print(f"\n  Done:{counter['done']:,} | Errors:{counter['errors']} | "
          f"Blocked:{counter['blocked']}")

    if not results:
        print("\n[WARN] Không crawl được sản phẩm nào!")
        print("  → Thử: HEADLESS = False  để xem browser hoạt động")
        sys.exit(1)

    nb = len(set(p["brand"] for p in results))
    np = sum(1 for p in results if p["list_price"] and p["list_price"] != "Quote only")
    print(f"\n[STEP 3] Exporting {len(results):,} products...")
    write_excel(results)

    print(f"\n{'='*65}")
    print(f"    Done!  Products:{len(results):,}  Brands:{nb}  "
          f"With price:{np:,}  Quote only:{len(results)-np:,}")
    print(f"    Output: {OUTPUT_FILE}")
    print(f"{'='*65}")

    if not stop_flag and MAX_PRODUCTS == 0:
        Path(CHECKPOINT).unlink(missing_ok=True)


H_FILL = PatternFill("solid", start_color="1F4E79")
H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ALT    = PatternFill("solid", start_color="D6E4F0")
WHT    = PatternFill("solid", start_color="FFFFFF")
SUM    = PatternFill("solid", start_color="BDD7EE")
NF     = Font(name="Arial", size=10)
LF_    = Font(name="Arial", size=10, color="0563C1", underline="single")
BF     = Font(name="Arial", bold=True, size=10)
GF     = Font(name="Arial", bold=True, size=10, color="1F6B1F")
CTR    = Alignment(horizontal="center", vertical="center")
LFT    = Alignment(horizontal="left",   vertical="center")
THIN   = Side(border_style="thin", color="BDD7EE")
BRD    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
COLS   = ["#","Brand","Model","Product Name","Category",
          "List Price","Your Price","In Stock","Description","URL"]
WIDTHS = [5,20,22,50,30,13,13,18,55,65]

def hdr(ws):
    for ci,(c,w) in enumerate(zip(COLS,WIDTHS),1):
        cell=ws.cell(1,ci,value=c)
        cell.font=H_FONT; cell.fill=H_FILL; cell.alignment=CTR; cell.border=BRD
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[1].height=20

def drow(ws, row, vals, lc=None):
    fill=ALT if row%2==0 else WHT
    for ci,val in enumerate(vals,1):
        c=ws.cell(row,ci,value=val)
        c.fill=fill; c.border=BRD; c.alignment=LFT
        if ci==lc and val and str(val).startswith("http"):
            c.font=LF_; c.hyperlink=str(val)
        elif ci==7 and val and str(val).startswith("$"):
            c.font=GF
        else:
            c.font=NF
    ws.row_dimensions[row].height=15

def safe_ws(s): return re.sub(r'[\\/*?\[\]:]','',s)[:31]

def write_excel(products: list):
    wb    = Workbook()
    prods = sorted(products, key=lambda p:(p["brand"].lower(), p["model"].lower()))

    ws=wb.active; ws.title="ALL_PRODUCTS"
    ws.sheet_view.showGridLines=False; ws.freeze_panes="A2"
    hdr(ws); ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}1"
    for i,p in enumerate(prods,2):
        drow(ws,i,[i-1,p["brand"],p["model"],p["name"],p["category"],
                   p["list_price"],p["your_price"],p["in_stock"],
                   (p["description"] or "")[:150],p["url"]],lc=10)
    print(f"  Sheet 'ALL_PRODUCTS': {len(prods):,} rows")

    ws2=wb.create_sheet("SUMMARY"); ws2.sheet_view.showGridLines=False
    ws2.merge_cells("A1:G1"); t=ws2["A1"]
    t.value="TEquipment.net — Price List Export"
    t.font=Font(name="Arial",bold=True,size=14,color="1F4E79"); t.alignment=CTR
    ws2.row_dimensions[1].height=28
    ws2.merge_cells("A2:G2"); d=ws2["A2"]
    nb=len(set(p["brand"] for p in prods))
    np=sum(1 for p in prods if p["list_price"] and p["list_price"]!="Quote only")
    d.value=(f"Crawled:{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
             f"{len(prods):,} products | {nb} brands | {np:,} priced | "
             f"{len(prods)-np:,} quote only")
    d.font=Font(name="Arial",italic=True,size=9,color="595959"); d.alignment=CTR

    SC=["#","Brand","Total","List Price","Your Price","Quote Only","Avg List"]
    SW=[5,28,10,12,12,12,14]
    for ci,(c,w) in enumerate(zip(SC,SW),1):
        cell=ws2.cell(4,ci,value=c)
        cell.font=H_FONT; cell.fill=H_FILL; cell.alignment=CTR; cell.border=BRD
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.auto_filter.ref="A4:G4"; ws2.freeze_panes="A5"

    bg=defaultdict(list)
    for p in prods: bg[p["brand"]].append(p)

    for i,(brand,bp) in enumerate(sorted(bg.items()),1):
        lv,yv,qo=[],[],0
        for p in bp:
            if p["list_price"]=="Quote only": qo+=1
            else:
                try: lv.append(float((p["list_price"] or "").replace("$","").replace(",","")))
                except: pass
                try: yv.append(float((p["your_price"] or "").replace("$","").replace(",","")))
                except: pass
        avg=f"${sum(lv)/len(lv):,.2f}" if lv else "N/A"
        drow(ws2,4+i,[i,brand,len(bp),len(lv),len(yv),qo,avg])

    tr=4+len(bg)+1
    tl=sum(1 for p in prods if p["list_price"] and p["list_price"]!="Quote only")
    ty=sum(1 for p in prods if p["your_price"])
    tq=sum(1 for p in prods if p["list_price"]=="Quote only")
    for ci,val in enumerate(["","TOTAL",len(prods),tl,ty,tq,""],1):
        c=ws2.cell(tr,ci,value=val)
        c.font=BF; c.fill=SUM; c.border=BRD; c.alignment=CTR

    for brand,bp in sorted(bg.items()):
        ws_b=wb.create_sheet(safe_ws(brand))
        ws_b.sheet_view.showGridLines=False; ws_b.freeze_panes="A2"
        hdr(ws_b); ws_b.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}1"
        for ri,p in enumerate(bp,2):
            drow(ws_b,ri,[ri-1,p["brand"],p["model"],p["name"],p["category"],
                          p["list_price"],p["your_price"],p["in_stock"],
                          (p["description"] or "")[:150],p["url"]],lc=10)
        print(f"  Sheet '{safe_ws(brand)}': {len(bp)} rows")

    wb.save(OUTPUT_FILE)
    print(f"\n  Saved → {OUTPUT_FILE}")


if __name__ == "__main__":
    asyncio.run(main_async())